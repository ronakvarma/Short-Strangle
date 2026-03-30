"""
BankNifty 09:20 AM Short Strangle Backtesting System
=====================================================
Strategy
--------
- Enter at 09:20 AM by shorting 1 CE + 1 PE (1 lot = 15 units) each Week-1 day
- Strike selection: closest 09:20 close premium to Rs.50 per leg
- Exit at 15:20 close, OR when 50% SL is breached (checked via High column)
- Only Week-1 calendar days (day 1-7); Wednesday = expiry day
- No compounding; no lookahead bias; fully vectorised
"""

import os
import sys
import time
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

warnings.filterwarnings("ignore")

# ======================================================================
# CONFIG
# ======================================================================
LOT_SIZE        = 15
TARGET_PREMIUM  = 50.0
ENTRY_TINT      = 920     # 09:20 as integer HHMM  (fastest filter)
EXIT_TINT       = 1520    # 15:20
ENTRY_TIME_STR  = "09:20"
EXIT_TIME_STR   = "15:20"
SL_PCT          = 0.50
INITIAL_NAV     = 100.0
INITIAL_CAPITAL = 100_000.0   # Rs. 1,00,000 starting capital

# ======================================================================
# MODULE 1  DATA LOADING
# ======================================================================

def load_data(path: str) -> pd.DataFrame:
    """Load 1-minute OHLCV data, normalise columns, parse datetime."""
    df = pd.read_csv(path, low_memory=False)
    df.columns = df.columns.str.strip()

    rename = {}
    for c in df.columns:
        cl = c.strip().lower()
        if   cl in ("ticker", "symbol"):  rename[c] = "Ticker"
        elif cl == "date":                 rename[c] = "Date"
        elif cl == "time":                 rename[c] = "Time"
        elif cl == "open":                 rename[c] = "Open"
        elif cl == "high":                 rename[c] = "High"
        elif cl == "low":                  rename[c] = "Low"
        elif cl == "close":                rename[c] = "Close"
        elif cl in ("volume", "vol"):      rename[c] = "Volume"
    df.rename(columns=rename, inplace=True)

    df["DateTime"] = pd.to_datetime(
        df["Date"].astype(str) + " " + df["Time"].astype(str),
        dayfirst=True, errors="coerce"
    )
    df.dropna(subset=["DateTime"], inplace=True)

    df["Date"]    = df["DateTime"].dt.date
    # Integer HHMM is faster than strftime for filtering
    df["TimeInt"] = df["DateTime"].dt.hour * 100 + df["DateTime"].dt.minute
    df["TimeStr"] = df["TimeInt"].apply(lambda x: f"{x // 100:02d}:{x % 100:02d}")

    for col in ("Open", "High", "Low", "Close"):
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df.dropna(subset=["Close"], inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def parse_tickers_vectorized(ticker_series: pd.Series):
    """
    Vectorised ticker parser for BANKNIFTY{strike}{CE/PE} format.
    Returns (strike Series, option_type Series) — no Python loops.
    """
    s      = ticker_series.str.upper().str.strip()
    is_opt = s.str.startswith("BANKNIFTY") & (s != "BANKNIFTY")
    body   = s.str[9:]                         # strip "BANKNIFTY" prefix

    ce_mask = body.str.endswith("CE")
    pe_mask = body.str.endswith("PE")

    opt_type_arr = np.where(ce_mask, "CE", np.where(pe_mask, "PE", None))
    strike_str   = np.where(ce_mask | pe_mask, body.str[:-2], None)

    opt_type = pd.Series(
        np.where(is_opt, opt_type_arr, None), index=ticker_series.index, dtype=object
    )
    strike = pd.to_numeric(
        pd.Series(np.where(is_opt, strike_str, None), index=ticker_series.index),
        errors="coerce"
    )
    return strike, opt_type


# ======================================================================
# MODULE 2  WEEK-1 / EXPIRY CALENDAR
# ======================================================================

def build_calendar(option_dates) -> pd.DataFrame:
    """
    Week-1  = calendar day 1-7 of each month.
    Expiry  = first Wednesday (weekday==2) in Week-1 that exists in data.
    """
    cal = pd.DataFrame({"Date": pd.to_datetime(option_dates)})
    cal["Month"]   = cal["Date"].dt.to_period("M")
    cal["Day"]     = cal["Date"].dt.day
    cal["Weekday"] = cal["Date"].dt.weekday    # 0=Mon, 2=Wed

    cal["IsWeek1"] = cal["Day"] <= 7

    wed_w1     = cal[cal["IsWeek1"] & (cal["Weekday"] == 2)]
    expiry_map = wed_w1.groupby("Month")["Date"].min().rename("ExpiryDate")

    cal = cal.merge(expiry_map, on="Month", how="left")
    cal["IsExpiry"] = cal["Date"] == cal["ExpiryDate"]
    return cal.set_index("Date")[["IsWeek1", "IsExpiry"]]


# ======================================================================
# MODULE 3  STRIKE SELECTION (VECTORISED)
# ======================================================================

def select_strikes(entry_snap: pd.DataFrame) -> pd.DataFrame:
    """
    Select one CE and one PE per date whose premium is closest to
    TARGET_PREMIUM (Rs. 50).

    Input cols  : Date, Ticker, Strike, OptionType, Close
    Returns     : one CE + one PE row per date with 'Leg' column added.
    """
    df = entry_snap.copy()
    df["PremiumDiff"] = (df["Close"] - TARGET_PREMIUM).abs()

    df_sorted = df.sort_values(["Date", "OptionType", "PremiumDiff"])
    best = (df_sorted
            .groupby(["Date", "OptionType"], sort=False, as_index=False)
            .first())
    best.rename(columns={"OptionType": "Leg"}, inplace=True)
    best.drop(columns=["PremiumDiff"], inplace=True)
    return best


# ======================================================================
# MODULE 4  STOP-LOSS CHECK (FULLY VECTORISED)
# ======================================================================

def check_stoploss(df_options_w1: pd.DataFrame,
                   entry_df: pd.DataFrame) -> pd.DataFrame:
    """
    For each leg find the EARLIEST intraday bar (09:21 to 15:19) where
    High >= EntryPrice * (1 + SL_PCT).

    Since we are SHORT: SL fires when price moves UP against us.
    SL exit price = EntryPrice * 1.50  (buy-back price at trigger level).

    Adds columns: SL_Hit, SL_Time, SL_Price, ExitTime, ExitPrice.
    """
    intra = df_options_w1[
        (df_options_w1["TimeInt"] > ENTRY_TINT) &
        (df_options_w1["TimeInt"] < EXIT_TINT)
    ][["Date", "Ticker", "TimeInt", "TimeStr", "High"]].copy()

    # Composite key for O(1) lookup
    entry_df        = entry_df.copy()
    entry_df["_DT"] = entry_df["Date"].astype(str) + "|" + entry_df["Ticker"]
    intra["_DT"]    = intra["Date"].astype(str)    + "|" + intra["Ticker"]

    sl_map           = entry_df.set_index("_DT")["EntryPrice"] * (1 + SL_PCT)
    intra["SLLevel"] = intra["_DT"].map(sl_map)
    intra["Hit"]     = intra["High"] >= intra["SLLevel"]

    # First breached bar per trade key
    first = (
        intra[intra["Hit"]]
        .sort_values("TimeInt")
        .groupby("_DT", as_index=False)
        .first()[["_DT", "TimeStr", "SLLevel"]]
        .rename(columns={"TimeStr": "SL_Time", "SLLevel": "SL_Price"})
    )
    first["SL_Hit"] = True

    entry_df = entry_df.merge(first, on="_DT", how="left")
    entry_df["SL_Hit"]    = entry_df["SL_Hit"].fillna(False)
    entry_df["ExitTime"]  = np.where(entry_df["SL_Hit"],
                                     entry_df["SL_Time"],
                                     EXIT_TIME_STR)
    entry_df["ExitPrice"] = np.where(entry_df["SL_Hit"],
                                     entry_df["SL_Price"],
                                     entry_df["NormalExitPrice"])
    entry_df.drop(columns=["_DT"], inplace=True)
    return entry_df


# ======================================================================
# MODULE 5  STATISTICS
# ======================================================================

def compute_statistics(trade_df: pd.DataFrame,
                       daily_nav: pd.DataFrame) -> dict:
    stats = {}
    nav   = daily_nav["NAV"].values
    dates_sorted = pd.to_datetime(daily_nav["Date"].astype(str)).sort_values()
    n_years = (dates_sorted.iloc[-1] - dates_sorted.iloc[0]).days / 365.25

    # CAGR: (final_NAV / initial_NAV) ^ (1/n_years) - 1
    # NAV already starts at INITIAL_NAV=100, so this is correct directly.
    stats["CAGR"] = ((nav[-1] / INITIAL_NAV) ** (1 / n_years) - 1
                     if n_years > 0 else 0.0)

    # Max drawdown — calculated on the NAV series trade-by-trade
    roll_max  = np.maximum.accumulate(nav)
    dd_series = (nav - roll_max) / roll_max
    stats["MaxDrawdown"]    = float(dd_series.min())
    stats["DrawdownSeries"] = dd_series

    # Winners / Losers
    for leg in ("CE", "PE", "Combined"):
        sub   = trade_df if leg == "Combined" else trade_df[trade_df["Leg"] == leg]
        total = max(len(sub), 1)
        wins  = int((sub["GrossPnL"] > 0).sum())
        loss  = int((sub["GrossPnL"] <= 0).sum())
        stats[f"{leg}_Winners"] = wins
        stats[f"{leg}_Losers"]  = loss
        stats[f"{leg}_WinPct"]  = wins / total * 100
        stats[f"{leg}_LossPct"] = loss / total * 100

    # Avg % P&L — expiry vs non-expiry
    for leg in ("CE", "PE", "Combined"):
        for is_exp, tag in ((True, "Expiry"), (False, "NonExpiry")):
            sub = trade_df[trade_df["IsExpiry"] == is_exp]
            if leg != "Combined":
                sub = sub[sub["Leg"] == leg]
            pct = (sub["GrossPnL"] / sub["EntryValue"] * 100).mean() if len(sub) else 0.0
            stats[f"{leg}_{tag}_AvgPctPnL"] = float(pct)

    nav_df          = daily_nav.copy()
    nav_df["Date"]  = pd.to_datetime(nav_df["Date"].astype(str))
    nav_df["Month"] = nav_df["Date"].dt.to_period("M")
    monthly = (nav_df.groupby("Month")["NAV"]
               .agg(EndNAV="last")
               .reset_index())
    monthly["StartNAV"] = monthly["EndNAV"].shift(1).fillna(INITIAL_NAV)
    monthly["MonthlyPnLPct"] = (
        (monthly["EndNAV"] - monthly["StartNAV"]) / monthly["StartNAV"] * 100
    )
    stats["Monthly"] = monthly
    return stats


# ======================================================================
# MODULE 6  CHART GENERATION
# ======================================================================

def generate_charts(daily_nav: pd.DataFrame,
                    stats: dict, out_dir: str) -> tuple:
    dates = pd.to_datetime(daily_nav["Date"].astype(str))
    nav   = daily_nav["NAV"].values
    dd    = stats["DrawdownSeries"]

    # Equity curve
    fig, ax = plt.subplots(figsize=(14, 5))
    ax.plot(dates, nav, color="#1565C0", linewidth=1.8, label="NAV")
    ax.fill_between(dates, INITIAL_NAV, nav,
                    where=(nav >= INITIAL_NAV), alpha=0.12, color="#1565C0")
    ax.fill_between(dates, INITIAL_NAV, nav,
                    where=(nav < INITIAL_NAV),  alpha=0.12, color="#C62828")
    ax.axhline(INITIAL_NAV, color="grey", linestyle="--", linewidth=0.8)
    ax.set_title("Equity Curve  (Base NAV = 100)", fontsize=13, fontweight="bold")
    ax.set_ylabel("NAV"); ax.set_xlabel("Date")
    ax.grid(alpha=0.25); ax.legend(fontsize=9)
    plt.tight_layout()
    eq_path = os.path.join(out_dir, "equity_curve.png")
    fig.savefig(eq_path, dpi=150, bbox_inches="tight"); plt.close(fig)

    # Drawdown curve
    fig, ax = plt.subplots(figsize=(14, 4))
    ax.fill_between(dates, dd * 100, 0, alpha=0.55, color="#C62828")
    ax.plot(dates, dd * 100, color="#C62828", linewidth=0.9)
    ax.axhline(dd.min() * 100, color="#7B0000", linestyle="--", linewidth=0.9,
               label=f"Max DD  {dd.min()*100:.2f}%")
    ax.set_title("Drawdown from Peak Equity", fontsize=13, fontweight="bold")
    ax.set_ylabel("Drawdown (%)"); ax.set_xlabel("Date")
    ax.yaxis.set_major_formatter(mtick.PercentFormatter())
    ax.grid(alpha=0.25); ax.legend(fontsize=9)
    plt.tight_layout()
    dd_path = os.path.join(out_dir, "drawdown.png")
    fig.savefig(dd_path, dpi=150, bbox_inches="tight"); plt.close(fig)

    return eq_path, dd_path


# ======================================================================
# MODULE 7  EXCEL EXPORT
# ======================================================================

_NAVY   = "1F4E79"
_BLUE   = "2E75B6"
_LTBLUE = "D6E4F0"
_WHITE  = "FFFFFF"


def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_row(ws, row, n_cols, fill_hex=_NAVY):
    fill   = PatternFill("solid", fgColor=fill_hex)
    font   = Font(bold=True, color=_WHITE, name="Arial", size=10)
    border = _border()
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font
        cell.border = border; cell.alignment = center


def _cell(cell, value, fmt=None, bold=False, alt=False,
          align="right", color="000000"):
    cell.value     = value
    cell.font      = Font(bold=bold, name="Arial", size=9, color=color)
    cell.border    = _border()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if alt:
        cell.fill = PatternFill("solid", fgColor=_LTBLUE)
    if fmt:
        cell.number_format = fmt


# ---------- Guide sheet ----------
def build_guide(wb: Workbook, elapsed: float):
    ws = wb.active
    ws.title = "Guide"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 130

    def add(row, text, size=10, bold=False, color="000000",
            bg=None, center=False):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=bold, name="Arial", size=size, color=color)
        c.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=True)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[row].height = 24 if size >= 14 else 16

    r = 1
    add(r, "BankNifty 09:20 AM Short Strangle  Backtest Report",
        size=16, bold=True, color=_NAVY, bg=_LTBLUE, center=True); r += 2

    sections = [
        ("Objective", 12,
         "Develop and run a one-year backtest for a 09:20 AM Short Strangle "
         "strategy on BankNifty options using 1-minute OHLCV data."),
        ("Strategy Overview", 11,
         "Short one CE and one PE simultaneously at 09:20 AM on every Week-1 "
         "trading day (calendar days 1-7 of each month). "
         "Exit at 15:20 or when a 50% stop-loss is breached on any leg."),
        ("Strike Selection", 11,
         "At 09:20 AM, compare the 1-min close premium of all available CE and "
         "PE options. Select the strike whose premium is closest to Rs. 50 for "
         "each leg. Both legs may have different strike prices."),
        ("Stop-Loss Logic", 11,
         "SL trigger = Entry Price x 1.50. The High column is checked from "
         "09:21 up to (but NOT including) 15:20. Since we are short, SL fires "
         "when High >= trigger level. Exit price = trigger level (EntryPrice x 1.50). "
         "If no SL fires, exit at the 15:20 1-min close price."),
        ("Position Sizing", 11,
         "Fixed 1 lot per day. Lot size = 15 units. No compounding across days. "
         "Available Capital = INITIAL_CAPITAL + cumulative P&L."),
        ("Week-1 and Expiry", 11,
         "Only calendar days 1-7 of each month are traded. "
         "Expiry day = first Wednesday that falls in Week-1 with available data. "
         "The strategy is entered every Week-1 day regardless of expiry."),
        ("CAGR — How It Is Calculated", 11,
         "CAGR = (Final NAV / 100) ^ (1 / n_years) - 1, where n_years = "
         "(last_date - first_date).days / 365.25. "
         "This uses the actual calendar span of the backtest. "
         "DO NOT use trading-day count / 252 — Week-1 only has ~55-60 trading "
         "days per year, making n_years ≈ 0.22 and inflating CAGR by ~4.5x."),
        ("Monthly P&L Chain", 11,
         "StartNAV of each month = EndNAV of the prior month (not the first "
         "trade of the current month). First month StartNAV = 100 (INITIAL_NAV). "
         "This ensures monthly % returns sum correctly to the overall return."),
        ("Output Worksheets", 11,
         "1. Guide        - This documentation sheet.\n"
         "2. TradeSheet   - Row-by-row trade log with P&L% per trade.\n"
         "3. Statistics   - KPIs, NAV table, monthly P&L, embedded charts."),
    ]

    for title, sz, body in sections:
        add(r, title, size=sz, bold=True, color=_BLUE); r += 1
        add(r, body,  size=9,  color="333333"); r += 2

    add(r, f"Total code execution time: {elapsed:.2f} seconds "
        f"(target < 60s including all data processing)",
        size=9, color="595959"); r += 1
    add(r, "Code is fully vectorised. No Python-level loops over rows.",
        size=8, color="595959")


# ---------- TradeSheet ----------
TRADE_COLS = [
    ("Entry Date",        "Date",             None),
    ("Entry Time",        "_ENTRY",           None),
    ("Exit Date",         "Date",             None),
    ("Exit Time",         "ExitTime",         None),
    ("Option Ticker",     "Ticker",           None),
    ("Strike Price",      "Strike",           "#,##0"),
    ("Option Type",       "Leg",              None),
    ("Entry Price",       "EntryPrice",       "#,##0.00"),
    ("Exit Price",        "ExitPrice",        "#,##0.00"),
    ("Quantity",          "Quantity",         "#,##0"),
    ("Entry Value",       "EntryValue",       "#,##0.00"),
    ("Exit Value",        "ExitValue",        "#,##0.00"),
    ("Gross P&L",         "GrossPnL",         "#,##0.00"),
    ("P&L %",             "PnLPct",           "0.00%"),   
    ("Cumulative P&L",    "CumPnL",           "#,##0.00"),
    ("Available Capital", "AvailableCapital", "#,##0.00"),
    ("BankNifty Close",   "UnderlyingClose",  "#,##0.00"),
    ("Is Expiry",         "IsExpiry",         None),
    ("SL Hit",            "SL_Hit",           None),
]


def build_tradesheet(wb: Workbook, tdf: pd.DataFrame):
    ws = wb.create_sheet("TradeSheet")
    ws.sheet_view.showGridLines = False

    headers = [c[0] for c in TRADE_COLS]
    keys    = [c[1] for c in TRADE_COLS]
    fmts    = [c[2] for c in TRADE_COLS]

    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _hdr_row(ws, 1, len(headers))
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    records = tdf.to_dict("records")
    for ri, row in enumerate(records, 2):
        alt = (ri % 2 == 0)
        for ci, (key, fmt) in enumerate(zip(keys, fmts), 1):
            if key == "_ENTRY":
                val = ENTRY_TIME_STR
            else:
                val = row.get(key, "")
                if isinstance(val, (np.bool_, bool)):
                    val = bool(val)
                elif isinstance(val, float) and np.isnan(val):
                    val = ""
                elif key == "PnLPct" and isinstance(val, (int, float)):
                    val = val / 100.0
            _cell(ws.cell(row=ri, column=ci), val, fmt, alt=alt)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ---------- Statistics sheet ----------
def build_stats(wb: Workbook, stats: dict, daily_nav: pd.DataFrame,
                eq_img: str, dd_img: str):
    ws = wb.create_sheet("Statistics")
    ws.sheet_view.showGridLines = False
    for ci, w in enumerate([24, 16, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    r = 1

    def sec(text, row, n=5):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=n)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = PatternFill("solid", fgColor=_BLUE)
        c.font = Font(bold=True, color=_WHITE, name="Arial", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 18
        return row + 1

    def kv(key, val, row, fmt=None):
        kc = ws.cell(row=row, column=1, value=key)
        kc.font = Font(bold=True, name="Arial", size=10)
        kc.border = _border()
        kc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=row, column=2, value=val)
        vc.font = Font(name="Arial", size=10)
        vc.border = _border()
        vc.alignment = Alignment(horizontal="right", vertical="center")
        if fmt: vc.number_format = fmt
        return row + 1

    def sub_hdr(ws, row, labels):
        for ci, h in enumerate(labels, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = PatternFill("solid", fgColor=_NAVY)
            c.font = Font(bold=True, color=_WHITE, name="Arial", size=10)
            c.border = _border()
            c.alignment = Alignment(horizontal="center", vertical="center")
        return row + 1

    # Key metrics
    r = sec("Key Performance Metrics", r)
    r = kv("CAGR",         stats["CAGR"],       r, "0.00%")
    r = kv("Max Drawdown", stats["MaxDrawdown"], r, "0.00%")
    r += 1

    # Winners / Losers
    r = sec("Winners and Losers", r)
    r = sub_hdr(ws, r, ["Metric", "CE", "PE", "Combined"])
    for label, key, fmt in [
        ("Winners", "Winners",  None),
        ("Losers",  "Losers",   None),
        ("Win %",   "WinPct",   "0.00%"),
        ("Loss %",  "LossPct",  "0.00%"),
    ]:
        alt = (r % 2 == 0)
        _cell(ws.cell(row=r, column=1), label, bold=True, align="left")
        for ci, leg in enumerate(["CE", "PE", "Combined"], 2):
            raw = stats.get(f"{leg}_{key}", 0)
            val = raw / 100 if fmt == "0.00%" else raw
            _cell(ws.cell(row=r, column=ci), val, fmt, alt=alt)
        r += 1
    r += 1

    # Avg % P&L
    r = sec("Average Percent P&L by Day Type", r)
    r = sub_hdr(ws, r, ["Period", "CE", "PE", "Combined"])
    for tag, label in [("Expiry", "Expiry Days"), ("NonExpiry", "Non-Expiry Days")]:
        alt = (r % 2 == 0)
        _cell(ws.cell(row=r, column=1), label, bold=True, align="left")
        for ci, leg in enumerate(["CE", "PE", "Combined"], 2):
            val = stats.get(f"{leg}_{tag}_AvgPctPnL", 0) / 100
            _cell(ws.cell(row=r, column=ci), val, "0.00%", alt=alt)
        r += 1
    r += 1

    # Monthly P&L
    r = sec("Monthly Percent P&L from NAV", r)
    r = sub_hdr(ws, r, ["Month", "Start NAV", "End NAV", "Monthly P&L %"])
    for mi, mrow in stats["Monthly"].iterrows():
        alt = (r % 2 == 0)
        _cell(ws.cell(row=r, column=1), str(mrow["Month"]), align="center")
        _cell(ws.cell(row=r, column=2), mrow["StartNAV"], "#,##0.00", alt=alt)
        _cell(ws.cell(row=r, column=3), mrow["EndNAV"],   "#,##0.00", alt=alt)
        _cell(ws.cell(row=r, column=4),
              mrow["MonthlyPnLPct"] / 100, "0.00%", alt=alt)
        r += 1
    r += 1

    # Daily NAV
    r = sec("Daily NAV Equity Curve", r)
    r = sub_hdr(ws, r, ["Date", "Daily P&L", "NAV"])
    for ni, nrow in daily_nav.iterrows():
        alt = (r % 2 == 0)
        _cell(ws.cell(row=r, column=1), str(nrow["Date"]), align="center")
        _cell(ws.cell(row=r, column=2), nrow["DailyPnL"], "#,##0.00", alt=alt)
        _cell(ws.cell(row=r, column=3), nrow["NAV"],      "#,##0.00", alt=alt)
        r += 1

    # Embed chart images
    chart_col = get_column_letter(7)
    if os.path.exists(eq_img):
        img = XLImage(eq_img)
        img.width = 700; img.height = 260
        ws.add_image(img, f"{chart_col}2")
    if os.path.exists(dd_img):
        img2 = XLImage(dd_img)
        img2.width = 700; img2.height = 210
        ws.add_image(img2, f"{chart_col}20")

    ws.freeze_panes = "A2"


# ======================================================================
# MAIN BACKTEST ENGINE
# ======================================================================

def run_backtest(data_path: str):
    t0 = time.perf_counter()

    print("[1/7] Loading and parsing data ...")
    df = load_data(data_path)
    df["Strike"], df["OptionType"] = parse_tickers_vectorized(df["Ticker"])

    df_index   = df[df["OptionType"].isna()].copy()
    df_options = df[df["OptionType"].notna()].copy()
    print(f"      Total rows: {len(df):,}  | "
          f"Options: {len(df_options):,}  | "
          f"Index: {len(df_index):,}")

    print("[2/7] Building Week-1 / expiry calendar ...")
    opt_dates    = pd.to_datetime(df_options["Date"].astype(str).unique())
    cal          = build_calendar(opt_dates)
    week1_dates  = set(cal[cal["IsWeek1"]].index.normalize())
    expiry_dates = set(cal[cal["IsExpiry"]].index.normalize())

    opt_date_dt   = pd.to_datetime(df_options["Date"].astype(str))
    df_options_w1 = df_options[opt_date_dt.isin(week1_dates)].copy()
    print(f"      Week-1 trading days: {len(week1_dates)}  | "
          f"Expiry days: {len(expiry_dates)}")

    print("[3/7] Extracting entry prices at 09:20 ...")
    entry_snap = df_options_w1[df_options_w1["TimeInt"] == ENTRY_TINT][
        ["Date", "Ticker", "Strike", "OptionType", "Close"]
    ].copy()

    print("[4/7] Selecting strikes (closest premium to Rs. 50) ...")
    selected = select_strikes(entry_snap)
    selected.rename(columns={"Close": "EntryPrice"}, inplace=True)

    print("[5/7] Extracting exit prices at 15:20 ...")
    exit_snap = (df_options_w1[df_options_w1["TimeInt"] == EXIT_TINT]
                 [["Date", "Ticker", "Close"]]
                 .copy()
                 .rename(columns={"Close": "NormalExitPrice"}))
    selected = selected.merge(exit_snap, on=["Date", "Ticker"], how="left")

    idx_entry = (df_index[df_index["TimeInt"] == ENTRY_TINT]
                 [["Date", "Close"]]
                 .rename(columns={"Close": "UnderlyingClose"}))
    selected = selected.merge(idx_entry, on="Date", how="left")

    print("[6/7] Vectorised stop-loss check ...")
    selected = check_stoploss(df_options_w1, selected)

    print("[7/7] Computing P&L, NAV and statistics ...")
    selected["Quantity"]   = LOT_SIZE
    selected["EntryValue"] = selected["EntryPrice"] * LOT_SIZE
    selected["ExitValue"]  = selected["ExitPrice"]  * LOT_SIZE
    # SHORT: profit = sell high, buy back low
    selected["GrossPnL"]   = (selected["EntryPrice"] -
                               selected["ExitPrice"]) * LOT_SIZE

    selected["PnLPct"] = (
        selected["GrossPnL"] / selected["EntryValue"] * 100
    ).round(4)

    sel_date_dt          = pd.to_datetime(selected["Date"].astype(str))
    selected["IsExpiry"] = sel_date_dt.isin(expiry_dates)

    selected.sort_values(["Date", "Leg"], inplace=True, ignore_index=True)
    selected["CumPnL"] = selected["GrossPnL"].cumsum()

    daily = (selected.groupby("Date", sort=True)["GrossPnL"]
             .sum().reset_index()
             .rename(columns={"GrossPnL": "DailyPnL"}))

    daily["AvailableCapital"] = INITIAL_CAPITAL + daily["DailyPnL"].cumsum()

    # NAV (base-100 index): NAV = 100 * (AvailableCapital / INITIAL_CAPITAL)
    daily["NAV"] = INITIAL_NAV * (daily["AvailableCapital"] / INITIAL_CAPITAL)

    selected = selected.merge(daily[["Date", "AvailableCapital"]],
                              on="Date", how="left")

    stats   = compute_statistics(selected, daily)
    elapsed = time.perf_counter() - t0
    print(f"\nBacktest complete in {elapsed:.2f}s")
    return selected, daily, stats, elapsed


# ======================================================================
# EXPORT WRAPPER
# ======================================================================

def export_excel(trade_df, daily_nav, stats, elapsed, out_path):
    out_dir = os.path.dirname(out_path) or "."
    os.makedirs(out_dir, exist_ok=True)

    eq_img, dd_img = generate_charts(daily_nav, stats, out_dir)

    wb = Workbook()
    build_guide(wb, elapsed)
    build_tradesheet(wb, trade_df)
    build_stats(wb, stats, daily_nav, eq_img, dd_img)

    wb.save(out_path)
    print(f"Excel saved -> {out_path}")
    return out_path


# ======================================================================
# ENTRY POINT
# ======================================================================

def main():
    if len(sys.argv) > 1:
        data_path = sys.argv[1]
    else:
        # Default: look for the largest CSV in the same folder as this script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        candidates = sorted(
            [f for f in os.listdir(script_dir) if f.lower().endswith(".csv")],
            key=lambda f: os.path.getsize(os.path.join(script_dir, f)),
            reverse=True
        )

        if not candidates:
            print("ERROR: No CSV found. Pass the path as argument:")
            print("  python banknifty_backtest.py <path_to_data.csv>")
            sys.exit(1)

        data_path = os.path.join(script_dir, candidates[0])
        print(f"Auto-detected data file: {data_path}")

    out_dir  = os.path.dirname(os.path.abspath(__file__))
    out_xlsx = os.path.join(out_dir, "banknifty_strangle_backtest.xlsx")

    trade_df, daily_nav, stats, elapsed = run_backtest(data_path)

    total_pnl = daily_nav["DailyPnL"].sum()
    total_pct = total_pnl / INITIAL_CAPITAL * 100

    print("\n" + "=" * 58)
    print("  BACKTEST SUMMARY")
    print("=" * 58)
    print(f"  Total trades         : {len(trade_df)}")
    print(f"  Total P&L (Rs.)      : {total_pnl:,.2f}")
    print(f"  Total Return (%)     : {total_pct:.2f}%")
    print(f"  CAGR                 : {stats['CAGR']*100:.2f}%")
    print(f"  Max Drawdown         : {stats['MaxDrawdown']*100:.2f}%")
    print(f"  CE  Win / Loss       : {stats['CE_Winners']} / "
          f"{stats['CE_Losers']}  ({stats['CE_WinPct']:.1f}%)")
    print(f"  PE  Win / Loss       : {stats['PE_Winners']} / "
          f"{stats['PE_Losers']}  ({stats['PE_WinPct']:.1f}%)")
    print(f"  Combined Win %       : {stats['Combined_WinPct']:.1f}%")
    print(f"  Execution time       : {elapsed:.2f}s")
    print("=" * 58)

    export_excel(trade_df, daily_nav, stats, elapsed, out_xlsx)


if __name__ == "__main__":
    main()
